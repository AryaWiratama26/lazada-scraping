import time
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
import csv
from typing import Literal
from openpyxl import Workbook, load_workbook
import os

TYPE_FILE = Literal['csv', 'txt', 'xlsx']

CHROME_DRIVER_PATH = "/usr/bin/chromedriver"

OP = webdriver.ChromeOptions()
# OP.add_argument('--headless')  
OP.add_argument("--disable-blink-features=AutomationControlled")

OP.add_argument("--no-sandbox")
OP.add_argument("--disable-dev-shm-usage")
OP.add_argument("user-agent=Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/139.0.0.0 Safari/537.36")

service = Service(executable_path=CHROME_DRIVER_PATH)

driver = webdriver.Chrome(service=service, options=OP)



class LazadaBot:
    def __init__(self):
        pass
    


    def __save_to_file(self, file_type, price, title, sold, output_file):
        if file_type == 'csv':

            with open(f"{output_file}.csv", "a") as f:

                writer = csv.writer(f)
                writer.writerow([title, price, sold])
        elif file_type == 'txt':

            with open(f"{output_file}.txt", "a", encoding="utf-8") as f:
                f.write(f"{title}, {price}, {sold}\n")
        elif file_type == 'xlsx':          

            if os.path.exists(f"{output_file}.xlsx"):

                wb = load_workbook(f"{output_file}.xlsx")
                ws = wb.active
                ws.append([title, price, sold])
                wb.save(f"{output_file}.xlsx")
            else:         
                wb = Workbook()
                ws = wb.active
                ws.append([title, price, sold])
                wb.save(f"{output_file}.xlsx")  
        else:

            raise ValueError("Harus csv atau txt!")

    def scrap(self, keyword, n_data, output_file, file_type: TYPE_FILE = 'csv'):

        try:
            driver.get("https://www.lazada.co.id/")


            time.sleep(2)

            search_keyword = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "input[placeholder='Cari di Lazada']"))
            )

            search_keyword.send_keys(keyword)
            search_keyword.send_keys(Keys.ENTER)

            print("Start Scraping.....")

            

            scraped = 0

            while scraped < n_data:

                time.sleep(3)
                driver.execute_script(f"window.scrollTo(0, 1500);")
                time.sleep(2)

                div_card = WebDriverWait(driver, 20).until(
                        EC.presence_of_all_elements_located((By.CSS_SELECTOR, "div[data-qa-locator='product-item']"))
                )


                
                for card in div_card:


                    if scraped >= n_data:
                        print(f"Data berhasil di scrape: {scraped}")
                        break


                    price = card.find_element(By.XPATH, ".//span[contains(text(), 'Rp')]").text
                    title = card.find_element(By.CSS_SELECTOR, "a[title]").text
                

                    try:
                        sold = card.find_element(By.XPATH, ".//span[contains(text(), 'Terjual')]").text
                        sold = sold.replace("Terjual", "")
                        sold = sold.replace(" ", "")
                        sold = sold.replace(".", "")
                        sold = sold.replace(",", "")
                    except:
                        sold = 0

            

                    price = price.replace("Rp", "")
                    price = price.replace(".", "")
                    price = price.replace(",", "")                              
                
                    
                    self.__save_to_file(file_type, price, title, sold, output_file)

                    scraped += 1

                    print(f"{scraped}. Data di scraping, file bernama {output_file}.{file_type}")

                    
                
                try:
                    next_page = WebDriverWait(driver, 10).until(
                        EC.element_to_be_clickable((By.CSS_SELECTOR, "li[class='ant-pagination-next'][title='Next Page']"))
                    )


                    button_next = next_page.find_element(By.CSS_SELECTOR, "button[type='button']")
                    button_next.click()
                    print("Change to next pagee")
                    time.sleep(4)

                except:

                    break
            


            time.sleep(5)

        except Exception as e:
            
            print(e)

             